"""
Excel source connector.

extract()        — pandas path: loads the full workbook at once (fast, higher memory).
extract_chunks() — openpyxl read-only path: streams row-by-row (bounded memory).

Use extract() when the workbook fits comfortably in RAM.
Set source.excel.chunk_size > 0 in the profile to activate streaming.
openpyxl is already a transitive dependency of pandas[excel], so no new packages
are required.
"""

from collections.abc import Iterator

import openpyxl
import pandas as pd

from etl_framework.connectors.base import BaseSourceConnector
from etl_framework.logger import audit, get_logger
from etl_framework.models import RawSheet

logger = get_logger()


class ExcelConnector(BaseSourceConnector):
    def __init__(self, path: str) -> None:
        self._path = path

    @property
    def source_name(self) -> str:
        return self._path

    # ── Full-load path (chunk_size == 0) ─────────────────────────────────────

    def extract(self) -> list[RawSheet]:
        sheets: dict[str, pd.DataFrame] = pd.read_excel(
            self._path, sheet_name=None, dtype=str
        )
        result: list[RawSheet] = []
        for name, df in sheets.items():
            df.columns = df.columns.str.strip()
            if df.empty:
                audit("extract_skip", sheet=name, trigger="empty sheet")
                continue
            rows = df.to_dict(orient="records")
            result.append(RawSheet(name=name, rows=rows))
            audit("extract_complete", sheet=name, rows_total=len(rows))
        return result

    # ── Streaming path (chunk_size > 0) ──────────────────────────────────────

    def extract_chunks(self, chunk_size: int) -> Iterator[RawSheet]:
        """
        Stream the workbook row-by-row using openpyxl read-only mode.

        openpyxl read_only=True does not materialise the full sheet into memory;
        rows are fetched lazily via iter_rows(). This keeps peak memory at
        O(chunk_size) rather than O(total_rows).

        Each yielded RawSheet contains at most chunk_size rows. All cell values
        are coerced to str (or None) to match the pandas extract() contract so
        downstream sanitizer/transform/validation code is unaffected.

        The workbook is closed in a try/finally block — Python generators honour
        try/finally on both normal exhaustion and early abandonment (generator.close()).
        """
        wb = openpyxl.load_workbook(self._path, read_only=True, data_only=True)
        try:
            for ws in wb.worksheets:
                row_iter = ws.iter_rows(values_only=True)

                # First row is always the header
                try:
                    raw_headers = next(row_iter)
                except StopIteration:
                    audit("extract_skip", sheet=ws.title, trigger="empty sheet")
                    continue

                headers = [
                    str(h).strip() if h is not None else f"_col{i}"
                    for i, h in enumerate(raw_headers, 1)
                ]

                chunk: list[dict] = []
                total_rows = 0

                for raw_row in row_iter:
                    # Coerce every cell to str | None — mirrors pandas dtype=str behaviour
                    row = {
                        h: (str(v) if v is not None else None)
                        for h, v in zip(headers, raw_row)
                    }
                    chunk.append(row)

                    if len(chunk) >= chunk_size:
                        total_rows += len(chunk)
                        audit("extract_chunk", sheet=ws.title,
                              rows_total=len(chunk))
                        yield RawSheet(name=ws.title, rows=chunk)
                        chunk = []

                if chunk:
                    total_rows += len(chunk)
                    audit("extract_chunk", sheet=ws.title, rows_total=len(chunk))
                    yield RawSheet(name=ws.title, rows=chunk)

                if total_rows == 0:
                    audit("extract_skip", sheet=ws.title, trigger="empty sheet")
                else:
                    audit("extract_complete", sheet=ws.title, rows_total=total_rows)
        finally:
            wb.close()
